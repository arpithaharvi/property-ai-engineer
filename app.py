# ============================================================
# app.py — Property AI Engineer — Main Application
# Month 1 Final Version — includes SQLite database (Session 5)
# UPDATED: Universal property-type-aware extraction
#          (Residential, Flat, Commercial, IT_Park,
#           Vacant_Site, Industrial)
#
# Pages:
#   🏠 Home              — overview and instructions
#   📤 Upload Documents  — process PDFs (Week 3)
#   🔍 Document Search   — ask questions across documents (Week 2+3)
#   📊 Risk Analysis     — PDF + Excel risk scoring (Week 1 + Week 4)
#   📋 Reports           — full history from database + downloads
# ============================================================
# --- Streamlit Cloud fix: ChromaDB needs a newer SQLite than Debian ships ---
__import__('sqlite3')
import sys
sys.modules['sqlite3'] = sys.modules.pop('sqlite3')
import json
import os
import tempfile
from io import BytesIO

import pandas as pd
import streamlit as st

from config import UPLOADS_PATH, REPORTS_PATH
from modules.pdf_loader import process_pdf, split_into_chunks
from modules.extractor import analyze_property_row
from modules.rag import (
    build_vectorstore,
    clear_vectorstore,
    build_qa_chain
)
from modules.validator import validate_property_data
from modules.database import (
    init_db,
    save_document,
    save_property_with_analysis,
    save_excel_batch,
    get_all_documents,
    get_all_excel_batches
)
# --- Streamlit Cloud fix: bridge st.secrets to environment variables ---
try:
    if "GROQ_API_KEY" in st.secrets:
        os.environ["GROQ_API_KEY"] = st.secrets["GROQ_API_KEY"]
except Exception:
    pass  # running locally with .env — that's fine
    
# ── Page config — must be first Streamlit command ──
st.set_page_config(
    page_title="Property AI Engineer",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Initialize database on every app start ──
init_db()

# ── Ensure data folders exist ──
os.makedirs(UPLOADS_PATH, exist_ok=True)
os.makedirs(REPORTS_PATH, exist_ok=True)

# ============================================================
# SIDEBAR NAVIGATION
# ============================================================

st.sidebar.title("🏘️ Property AI")
st.sidebar.markdown("---")

page = st.sidebar.radio(
    "Navigation",
    options=[
        "🏠 Home",
        "📤 Upload Documents",
        "🔍 Document Search",
        "📊 Risk Analysis",
        "📋 Reports"
    ]
)

st.sidebar.markdown("---")
st.sidebar.markdown("**Stack:**")
st.sidebar.markdown("⚡ Groq Vision + LLaMA 3.3")
st.sidebar.markdown("🤗 HuggingFace Embeddings")
st.sidebar.markdown("🗄️ ChromaDB + SQLite")
st.sidebar.markdown("💰 Total cost: ₹0")

# ============================================================
# PAGE 1 — HOME
# ============================================================

if page == "🏠 Home":
    st.title("🏘️ Property AI Engineer")
    st.subheader(
        "AI-powered property document intelligence "
        "for banks and NBFCs"
    )

    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Documents Supported", "Digital + Scanned")
        st.metric("Languages", "English + Kannada")
    with col2:
        st.metric("Processing Time", "2-3 minutes")
        st.metric("Cost", "₹0")
    with col3:
        st.metric("Accuracy", "Bank-grade")
        st.metric("Property Types", "6 supported")

    st.markdown("---")
    st.subheader("📋 How to use this application")
    st.write("""
    **Step 1 — Upload Documents**
    Go to 📤 Upload Documents. Upload your property PDFs —
    digital or scanned, English or Kannada. The app processes
    them automatically and builds a searchable database.

    **Step 2 — Search Documents**
    Go to 🔍 Document Search. Ask any question about your
    uploaded documents. The app searches all documents and
    gives answers with source page references.

    **Step 3 — Risk Analysis**
    Go to 📊 Risk Analysis. Upload a PDF valuation report or
    an Excel property list to get AI-powered risk scores,
    investment scores, validation flags, and bank-ready status.
    Supports Residential, Flat, Commercial, IT Park, Vacant Site,
    and Industrial properties.

    **Step 4 — View Reports**
    Go to 📋 Reports to see full history of all analysed
    properties (persists between sessions) and download results.
    """)

    st.markdown("---")
    st.subheader("✅ Supported property types")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.write("🏠 Residential")
        st.write("🏢 Flat / Apartment")
    with col2:
        st.write("🏬 Commercial")
        st.write("💻 IT Park")
    with col3:
        st.write("🌳 Vacant Site")
        st.write("🏭 Industrial")

# ============================================================
# PAGE 2 — UPLOAD DOCUMENTS
# ============================================================

elif page == "📤 Upload Documents":
    st.title("📤 Upload Property Documents")
    st.write(
        "Upload valuation documents — digital or scanned, "
        "English or Kannada."
    )

    st.subheader("Step 1 — Select Property Case Type")
    case_type = st.radio(
        "What type of property are you valuing?",
        options=[
            "🏠 Single Property — one plot, standard valuation",
            "🏗️ Merged Property — multiple plots/owners, one building"
        ]
    )
    is_merged = "Merged" in case_type

    if is_merged:
        st.info(
            "📌 Merged Property: details extracted separately per "
            "owner/document, then combined totals shown for bank report."
        )
    else:
        st.info(
            "📌 Single Property: standard extraction from all documents."
        )

    st.subheader("Step 2 — Upload Files")
    uploaded_files = st.file_uploader(
        "Upload property documents",
        type=["pdf"],
        accept_multiple_files=True
    )

    if uploaded_files:
        st.info(
            f"📎 {len(uploaded_files)} file(s) selected: "
            f"{', '.join([f.name for f in uploaded_files])}"
        )

    if uploaded_files and st.button("🚀 Process All Documents"):

        clear_vectorstore()
        all_chunks = []
        processing_summary = []

        for uploaded_file in uploaded_files:
            st.write("---")
            st.write(f"📄 Processing: **{uploaded_file.name}**")

            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".pdf"
            ) as tmp:
                tmp.write(uploaded_file.read())
                tmp_path = tmp.name

            try:
                progress_bar = st.progress(0)
                status_text = st.empty()

                documents, digital_count, vision_count = process_pdf(
                    tmp_path,
                    uploaded_file.name,
                    progress_bar,
                    status_text
                )

                progress_bar.empty()
                status_text.empty()

                if not documents:
                    st.error(
                        f"❌ {uploaded_file.name} — "
                        f"No text could be extracted."
                    )
                    continue

                st.success(
                    f"✅ {uploaded_file.name} — "
                    f"{digital_count} digital + "
                    f"{vision_count} scanned = "
                    f"{len(documents)} pages"
                )

                languages = set(
                    doc.metadata.get("language") for doc in documents
                )
                lang_str = ', '.join(languages)
                st.write(f"   🌐 Languages: {lang_str}")

                file_chunks = split_into_chunks(documents)
                all_chunks.extend(file_chunks)

                try:
                    doc_id = save_document(
                        filename=uploaded_file.name,
                        total_pages=len(documents),
                        digital_pages=digital_count,
                        scanned_pages=vision_count,
                        languages=lang_str,
                        case_type="merged" if is_merged else "single"
                    )
                    st.write(f"   💾 Saved to database (ID: {doc_id})")
                except Exception as db_err:
                    st.warning(f"   ⚠️ Database save warning: {db_err}")
                    doc_id = None

                processing_summary.append({
                    "file": uploaded_file.name,
                    "pages": len(documents),
                    "digital": digital_count,
                    "vision": vision_count,
                    "chunks": len(file_chunks),
                    "languages": lang_str,
                    "db_id": doc_id
                })

            except Exception as e:
                st.error(
                    f"❌ Error processing {uploaded_file.name}: {str(e)}"
                )
            finally:
                os.unlink(tmp_path)

        if all_chunks:
            with st.spinner("🧠 Building vector database..."):
                vectorstore = build_vectorstore(all_chunks)

            st.divider()
            st.success(
                f"✅ {len(processing_summary)} documents indexed, "
                f"{vectorstore._collection.count()} chunks stored, "
                f"saved to database."
            )

            st.subheader("📊 Processing Summary")
            for s in processing_summary:
                st.write(
                    f"• **{s['file']}** — "
                    f"{s['pages']} pages, {s['chunks']} chunks, "
                    f"Language: {s['languages']}"
                )

            st.session_state["vectorstore"] = vectorstore
            st.session_state["processed_files"] = processing_summary
            st.session_state["is_merged"] = is_merged

            st.info(
                "✅ Documents ready. "
                "Go to 🔍 Document Search to ask questions."
            )
        else:
            st.error("❌ No text extracted from any uploaded file.")

# ============================================================
# PAGE 3 — DOCUMENT SEARCH
# ============================================================

elif page == "🔍 Document Search":
    st.title("🔍 Document Search")
    st.write("Ask questions across all uploaded property documents.")

    if "vectorstore" not in st.session_state:
        st.warning(
            "⚠️ No documents loaded. "
            "Please go to 📤 Upload Documents first."
        )
    else:
        vectorstore = st.session_state["vectorstore"]
        is_merged = st.session_state.get("is_merged", False)
        processed_files = st.session_state.get("processed_files", [])

        st.subheader("📚 Loaded Documents")
        for f in processed_files:
            st.write(f"• **{f['file']}** — {f['pages']} pages")

        st.divider()

        qa_chain, retriever = build_qa_chain(vectorstore, is_merged)

        if is_merged:
            st.subheader("💬 Merged Property Questions")
            preset_questions = [
                "List owner name and address from each document separately.",
                "List the area and survey number from each document separately.",
                "List the boundary schedule from each document separately.",
                "List the schedule of property from each document separately.",
                "What is the total combined area of all plots merged together?",
                "Give the combined outer boundary of the merged property.",
                "List all owners and their respective share in the property.",
                "Summarize all documents for a single bank valuation report.",
            ]
        else:
            st.subheader("💬 Property Questions")
            preset_questions = [
                "What is the market value of the property?",
                "What is the distress value?",
                "What valuation method was used?",
                "What is the recommended loan amount?",
                "What are the risk factors mentioned?",
                "What is the owner name and address?",
                "What is the property location and total area?",
                "Are there any legal or litigation issues mentioned?"
            ]

        st.write("**Quick questions:**")
        selected_question = None
        cols = st.columns(2)
        for i, question in enumerate(preset_questions):
            if cols[i % 2].button(question, key=f"q_{i}"):
                selected_question = question

        user_question = st.text_input(
            "Or type your own question:",
            value=selected_question if selected_question else "",
            placeholder="e.g. What are the boundaries of the property?"
        )
        submit = st.button("🔍 Get Answer", key="submit_btn")

        if user_question and (submit or selected_question):
            with st.spinner("🔍 Searching documents..."):
                answer = qa_chain.invoke(user_question)
                source_docs = retriever.invoke(user_question)

            st.subheader("🤖 Answer")
            st.success(answer)

            st.subheader("📍 Sources")
            sources_by_file = {}
            for doc in source_docs:
                fname = doc.metadata.get("source", "Unknown")
                page_num = doc.metadata.get("page", "unknown")
                method = doc.metadata.get("extraction_method", "")
                method_label = (
                    "Direct Read" if method == "direct_text"
                    else "Groq Vision"
                )
                if fname not in sources_by_file:
                    sources_by_file[fname] = []
                sources_by_file[fname].append({
                    "page": page_num,
                    "method": method_label,
                    "content": doc.page_content
                })

            for fname, pages in sources_by_file.items():
                st.write(f"📄 **{fname}**")
                for p in pages:
                    with st.expander(
                        f"Page {p['page']} — {p['method']}"
                    ):
                        st.write(p["content"])

# ============================================================
# PAGE 4 — RISK ANALYSIS
# ============================================================

elif page == "📊 Risk Analysis":
    st.title("📊 Property Risk Analysis")
    st.write(
        "Upload a valuation report PDF or an Excel property list "
        "to get AI-powered risk, investment, and validation analysis. "
        "Supports Residential, Flat, Commercial, IT Park, Vacant Site, "
        "and Industrial properties."
    )

    st.subheader("Step 1 — Choose Input Type")
    input_type = st.radio(
        "What would you like to analyse?",
        options=[
            "📄 PDF Valuation Report — single property, full analysis",
            "📊 Excel Property List — multiple properties, batch scoring"
        ]
    )
    is_pdf_mode = "PDF" in input_type

    st.divider()

    # ===========================================================
    # MODE 1 — PDF VALUATION REPORT
    # ===========================================================
    if is_pdf_mode:
        st.subheader("Step 2 — Upload Valuation Report PDF")
        st.write(
            "Upload any property valuation report — "
            "digital or scanned, English or Kannada. "
            "The AI will detect the property type automatically."
        )

        uploaded_pdf = st.file_uploader(
            "Upload valuation report PDF",
            type=["pdf"],
            key="risk_pdf"
        )

        if uploaded_pdf and st.button(
            "🚀 Extract and Analyse", key="analyse_pdf"
        ):
            from modules.llm import load_groq_client
            from config import GROQ_LLM_MODEL

            groq_client = load_groq_client()

            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".pdf"
            ) as tmp:
                tmp.write(uploaded_pdf.read())
                tmp_path = tmp.name

            try:
                st.write("📖 Reading document...")
                progress_bar = st.progress(0)
                status_text = st.empty()

                documents, digital_count, vision_count = process_pdf(
                    tmp_path,
                    uploaded_pdf.name,
                    progress_bar,
                    status_text
                )

                progress_bar.empty()
                status_text.empty()

                if not documents:
                    st.error("❌ Could not extract text from this PDF.")
                    st.stop()

                st.success(
                    f"✅ {len(documents)} pages read "
                    f"({digital_count} digital + "
                    f"{vision_count} scanned)"
                )

                languages = set(
                    doc.metadata.get("language") for doc in documents
                )
                try:
                    doc_id = save_document(
                        filename=uploaded_pdf.name,
                        total_pages=len(documents),
                        digital_pages=digital_count,
                        scanned_pages=vision_count,
                        languages=', '.join(languages),
                        case_type="single"
                    )
                except Exception:
                    doc_id = None

                st.write("🧠 Building search index...")
                chunks = split_into_chunks(documents)
                build_vectorstore(chunks)

                st.write("🔍 Detecting property type and extracting details...")

                # ============================================================
                # UPDATED EXTRACTION PROMPT — property-type-aware
                # Detects one of: Residential, Flat, Commercial, IT_Park,
                # Vacant_Site, Industrial — then extracts only the fields
                # relevant to that type. Irrelevant sections are filled
                # with "Not Applicable" (not a red flag), while missing-
                # but-relevant fields are "Not found" (IS a red flag,
                # caught by modules/validator.py).
                # ============================================================
                extraction_prompt = """
You are analysing a property valuation/legal document. This document
could describe ANY of these property types:
  - Residential (independent house/villa)
  - Flat (apartment/flat in a multi-unit building)
  - Commercial (office, retail, mixed-use building)
  - IT_Park (IT park, tech park, SEZ building)
  - Vacant_Site (empty plot/land, no construction)
  - Industrial (factory, warehouse, industrial shed)

STEP 1 — Read the document and determine which ONE of the six types
above best describes this property. Put your answer in "property_type"
using EXACTLY one of these strings: Residential, Flat, Commercial,
IT_Park, Vacant_Site, Industrial.

STEP 2 — Extract the COMMON fields below. If a common field is not
stated in the document, write "Not found".

STEP 3 — Extract ONLY the fields under the section matching your
STEP 1 answer. For the OTHER five sections (types that do NOT match),
fill every field in those sections with the exact string
"Not Applicable".

CRITICAL RULE:
  "Not found"      = applies to this property type, but document
                      did not state it.
  "Not Applicable" = does not apply to this property type at all.

Respond in JSON ONLY, with this exact structure and these exact keys:

{
  "property_type": "",

  "owner_name": "",
  "property_address": "",
  "survey_number": "",
  "total_area_sqft": "",
  "market_value": "",
  "guideline_value": "",
  "distress_value": "",
  "legal_status": "",
  "encumbrance_status": "",
  "litigation_status": "",
  "registration_date": "",
  "road_access": "",
  "approved_layout_plan": "",

  "residential_fields": {
    "building_age_years": "",
    "construction_type": "",
    "occupancy_profile": "",
    "boundary_compliance": "",
    "residual_economic_life": ""
  },

  "flat_fields": {
    "flat_number": "",
    "floor_number": "",
    "total_floors_in_building": "",
    "builder_developer_name": "",
    "society_apartment_name": "",
    "undivided_share_of_land_sqft": "",
    "carpet_area_sqft": "",
    "super_built_up_area_sqft": "",
    "maintenance_charges_monthly": "",
    "occupancy_certificate_status": ""
  },

  "commercial_fields": {
    "usage_permission_type": "",
    "lease_status": "",
    "tenant_details": "",
    "rental_yield_annual": "",
    "commercial_occupancy_certificate": "",
    "fire_safety_noc_status": "",
    "parking_provision": ""
  },

  "it_park_fields": {
    "sez_status": "",
    "floor_space_index_fsi_utilized": "",
    "power_load_sanctioned_kva": "",
    "backup_power_provision": "",
    "built_up_area_vs_land_area_ratio": "",
    "anchor_tenant_details": "",
    "it_park_certification_status": ""
  },

  "vacant_site_fields": {
    "zoning_classification": "",
    "conversion_status": "",
    "fencing_boundary_status": "",
    "site_topography": "",
    "adjacent_land_use": "",
    "development_potential_notes": ""
  },

  "industrial_fields": {
    "factory_license_status": "",
    "pollution_control_clearance": "",
    "zoning_classification": "",
    "power_load_sanctioned_kva": "",
    "effluent_treatment_status": "",
    "warehouse_godown_area_sqft": "",
    "machinery_fixtures_included": ""
  },

  "violations_observed": "",
  "recommended_loan": "",
  "valuation_method": "",
  "risk_factors": []
}
"""
                full_text = "\n\n".join(
                    doc.page_content for doc in documents
                )

                extraction_response = groq_client.chat.completions.create(
                    model=GROQ_LLM_MODEL,
                    messages=[
                        {
                            "role": "system",
                            "content": (
                                "You are a senior property valuation analyst "
                                "experienced across residential, commercial, "
                                "industrial, and land valuation. Extract "
                                "information and respond ONLY in valid JSON."
                            )
                        },
                        {
                            "role": "user",
                            "content": (
                                f"{extraction_prompt}\n\n"
                                f"Document:\n{full_text[:8000]}"
                            )
                        }
                    ],
                    temperature=0.0,
                    response_format={"type": "json_object"},
                    max_tokens=1800
                )

                extracted = json.loads(
                    extraction_response.choices[0].message.content
                )

                detected_type = extracted.get("property_type", "Unknown")
                st.info(f"🏷️ Detected property type: **{detected_type}**")

                # Validate extracted fields (Week 4, now type-aware)
                st.write("✅ Running validation checks...")
                validation = validate_property_data(extracted)

                # Score the extracted data
                st.write("📊 Calculating risk and investment scores...")

                scoring_response = groq_client.chat.completions.create(
                    model=GROQ_LLM_MODEL,
                    messages=[
                        {
                            "role": "system",
                            "content": (
                                "You are a senior bank credit analyst "
                                "experienced across all property types. "
                                "Respond in JSON only."
                            )
                        },
                        {
                            "role": "user",
                            "content": f"""
Analyse this property and respond with exactly these fields in JSON:
{{
  "risk_score": "Low Risk / Medium Risk / High Risk",
  "investment_score": <1-100>,
  "legal_risk": "Low / Medium / High",
  "market_liquidity": "High / Moderate / Low",
  "price_range": "<range>",
  "encumbrance_risk": "Clear / Needs Verification / High Risk",
  "recommendation": "APPROVE / CONDITIONAL APPROVE / REJECT",
  "confidence_score": <1-100>,
  "key_risks": [],
  "positive_factors": [],
  "reason_for_recommendation": "<explanation>"
}}

This is a {detected_type} property. Adjust your analysis
appropriately — for example, do not penalise a Vacant_Site for
having no building, and do not penalise a Flat for not having
independent boundary walls.

Property data: {json.dumps(extracted, indent=2)}
"""
                        }
                    ],
                    temperature=0.0,
                    response_format={"type": "json_object"},
                    max_tokens=1000
                )

                scores = json.loads(
                    scoring_response.choices[0].message.content
                )

                # ── DATABASE: save property + scores + validation ──
                try:
                    if doc_id:
                        prop_id = save_property_with_analysis(
                            doc_id, extracted, scores, validation
                        )
                        st.write(
                            f"💾 Analysis saved to database "
                            f"(Property ID: {prop_id})"
                        )
                except Exception as db_err:
                    st.warning(f"⚠️ Database save warning: {db_err}")

                # Display results
                st.divider()
                st.subheader("📊 Risk Analysis Results")

                recommendation = scores.get(
                    "recommendation", "CONDITIONAL APPROVE"
                )
                if recommendation == "APPROVE":
                    st.success(
                        f"✅ RECOMMENDATION: **{recommendation}**"
                    )
                elif recommendation == "REJECT":
                    st.error(
                        f"❌ RECOMMENDATION: **{recommendation}**"
                    )
                else:
                    st.warning(
                        f"⚠️ RECOMMENDATION: **{recommendation}**"
                    )

                st.write(
                    f"**Reason:** "
                    f"{scores.get('reason_for_recommendation', 'N/A')}"
                )

                col1, col2, col3, col4 = st.columns(4)
                col1.metric(
                    "Risk Score", scores.get("risk_score", "N/A")
                )
                col2.metric(
                    "Investment Score",
                    f"{scores.get('investment_score', 0)}/100"
                )
                col3.metric(
                    "Legal Risk", scores.get("legal_risk", "N/A")
                )
                col4.metric(
                    "Confidence",
                    f"{scores.get('confidence_score', 0)}%"
                )

                col1, col2 = st.columns(2)
                col1.metric(
                    "Market Liquidity",
                    scores.get("market_liquidity", "N/A")
                )
                col2.metric(
                    "Encumbrance",
                    scores.get("encumbrance_risk", "N/A")
                )

                # Validation Report
                st.divider()
                st.subheader("✅ Validation Report")

                if validation["is_bank_ready"]:
                    st.success(
                        "✅ Bank-ready — no critical fields missing."
                    )
                else:
                    st.error(
                        f"🔴 NOT bank-ready — "
                        f"{len(validation['critical_flags'])} "
                        f"critical issue(s) must be resolved first."
                    )

                if validation["critical_flags"]:
                    st.write("**🔴 Critical Issues (must fix):**")
                    for flag in validation["critical_flags"]:
                        st.write(
                            f"• **{flag['field']}** — {flag['message']}"
                        )

                if validation["warning_flags"]:
                    st.write("**🟡 Warnings (verify manually):**")
                    for flag in validation["warning_flags"]:
                        st.write(
                            f"• **{flag['field']}** — {flag['message']}"
                        )

                if validation["info_flags"]:
                    with st.expander("🔵 Additional Notes"):
                        for flag in validation["info_flags"]:
                            st.write(
                                f"• **{flag['field']}** — {flag['message']}"
                            )

                # ============================================================
                # Extracted details — UPDATED to show common fields plus
                # ONLY the type-specific section matching detected_type.
                # This replaces the old fixed residential-field display.
                # ============================================================
                st.divider()
                st.subheader("📋 Extracted Property Details")

                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**Property Type:** {detected_type}")
                    st.write(f"**Owner:** {extracted.get('owner_name', 'N/A')}")
                    st.write(f"**Address:** {extracted.get('property_address', 'N/A')}")
                    st.write(f"**Survey Number:** {extracted.get('survey_number', 'N/A')}")
                    st.write(f"**Total Area:** {extracted.get('total_area_sqft', 'N/A')} sq ft")
                    st.write(f"**Road Access:** {extracted.get('road_access', 'N/A')}")
                    st.write(f"**Legal Status:** {extracted.get('legal_status', 'N/A')}")

                with col2:
                    st.write(f"**Market Value:** {extracted.get('market_value', 'N/A')}")
                    st.write(f"**Distress Value:** {extracted.get('distress_value', 'N/A')}")
                    st.write(f"**Guideline Value:** {extracted.get('guideline_value', 'N/A')}")
                    st.write(f"**Recommended Loan:** {extracted.get('recommended_loan', 'N/A')}")
                    st.write(f"**Valuation Method:** {extracted.get('valuation_method', 'N/A')}")
                    st.write(f"**Registration Date:** {extracted.get('registration_date', 'N/A')}")
                    st.write(f"**Encumbrance Status:** {extracted.get('encumbrance_status', 'N/A')}")

                # Type-specific section — only show the section matching
                # the detected property type, since the other five are
                # entirely "Not Applicable" and not useful to display
                type_section_map = {
                    "Residential": "residential_fields",
                    "Flat": "flat_fields",
                    "Commercial": "commercial_fields",
                    "IT_Park": "it_park_fields",
                    "Vacant_Site": "vacant_site_fields",
                    "Industrial": "industrial_fields",
                }
                section_key = type_section_map.get(detected_type)

                if section_key and isinstance(extracted.get(section_key), dict):
                    st.divider()
                    st.subheader(f"🏷️ {detected_type}-Specific Details")
                    type_data = extracted[section_key]
                    cols = st.columns(2)
                    for i, (field_name, field_value) in enumerate(type_data.items()):
                        display_name = field_name.replace("_", " ").title()
                        cols[i % 2].write(f"**{display_name}:** {field_value}")

                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("⚠️ Risk Factors")
                    for risk in scores.get("key_risks", []):
                        st.write(f"• {risk}")
                with col2:
                    st.subheader("✅ Positive Factors")
                    for pos in scores.get("positive_factors", []):
                        st.write(f"• {pos}")

                # Save to session state
                st.session_state["pdf_risk_result"] = {
                    "extracted": extracted,
                    "scores": scores,
                    "validation": validation,
                    "filename": uploaded_pdf.name
                }

                # ============================================================
                # STRUCTURED EXCEL EXPORT — Field / Value rows, grouped into
                # labeled sections, instead of one wide unreadable row.
                # Loan officers read this top-to-bottom like a report, not
                # side-to-side like raw data.
                # ============================================================

                def _fmt(value):
                    # Lists/dicts render awkwardly in Excel cells — flatten
                    # them to a readable comma-joined string
                    if isinstance(value, list):
                        return ", ".join(str(v) for v in value) if value else "None"
                    if isinstance(value, dict):
                        return "; ".join(f"{k}: {v}" for k, v in value.items())
                    return value

                report_rows = []  # list of (Section, Field, Value) tuples

                # ── Section 1: Property Identification ──
                common_fields = [
                    ("Property Type", detected_type),
                    ("Owner Name", extracted.get("owner_name")),
                    ("Property Address", extracted.get("property_address")),
                    ("Survey Number", extracted.get("survey_number")),
                    ("Total Area (sq ft)", extracted.get("total_area_sqft")),
                    ("Road Access", extracted.get("road_access")),
                    ("Approved Layout Plan", extracted.get("approved_layout_plan")),
                    ("Registration Date", extracted.get("registration_date")),
                ]
                for field, value in common_fields:
                    report_rows.append(("Property Identification", field, _fmt(value)))

                # ── Section 2: Legal & Compliance ──
                legal_fields = [
                    ("Legal Status", extracted.get("legal_status")),
                    ("Encumbrance Status", extracted.get("encumbrance_status")),
                    ("Litigation Status", extracted.get("litigation_status")),
                    ("Violations Observed", extracted.get("violations_observed")),
                ]
                for field, value in legal_fields:
                    report_rows.append(("Legal & Compliance", field, _fmt(value)))

                # ── Section 3: Valuation ──
                valuation_fields = [
                    ("Market Value", extracted.get("market_value")),
                    ("Guideline Value", extracted.get("guideline_value")),
                    ("Distress Value", extracted.get("distress_value")),
                    ("Recommended Loan", extracted.get("recommended_loan")),
                    ("Valuation Method", extracted.get("valuation_method")),
                ]
                for field, value in valuation_fields:
                    report_rows.append(("Valuation", field, _fmt(value)))

                # ── Section 4: Type-specific details (only the matching type) ──
                if section_key and isinstance(extracted.get(section_key), dict):
                    section_label = f"{detected_type}-Specific Details"
                    for field_name, field_value in extracted[section_key].items():
                        display_name = field_name.replace("_", " ").title()
                        report_rows.append((section_label, display_name, _fmt(field_value)))

                # ── Section 5: AI Risk & Investment Scores ──
                score_fields = [
                    ("Risk Score", scores.get("risk_score")),
                    ("Investment Score", scores.get("investment_score")),
                    ("Legal Risk", scores.get("legal_risk")),
                    ("Market Liquidity", scores.get("market_liquidity")),
                    ("Price Range", scores.get("price_range")),
                    ("Encumbrance Risk", scores.get("encumbrance_risk")),
                    ("Recommendation", scores.get("recommendation")),
                    ("Confidence Score (%)", scores.get("confidence_score")),
                    ("Reason for Recommendation", scores.get("reason_for_recommendation")),
                    ("Key Risks", scores.get("key_risks")),
                    ("Positive Factors", scores.get("positive_factors")),
                ]
                for field, value in score_fields:
                    report_rows.append(("AI Risk & Investment Analysis", field, _fmt(value)))

                # ── Section 6: Validation / Bank Readiness ──
                report_rows.append((
                    "Validation & Bank Readiness", "Bank Ready",
                    "Yes" if validation["is_bank_ready"] else "No"
                ))
                if validation["critical_flags"]:
                    for f in validation["critical_flags"]:
                        report_rows.append((
                            "Validation & Bank Readiness",
                            f"CRITICAL — {f['field']}", f["message"]
                        ))
                else:
                    report_rows.append((
                        "Validation & Bank Readiness", "Critical Issues", "None"
                    ))
                if validation["warning_flags"]:
                    for f in validation["warning_flags"]:
                        report_rows.append((
                            "Validation & Bank Readiness",
                            f"WARNING — {f['field']}", f["message"]
                        ))
                if validation["info_flags"]:
                    for f in validation["info_flags"]:
                        report_rows.append((
                            "Validation & Bank Readiness",
                            f"INFO — {f['field']}", f["message"]
                        ))

                report_df = pd.DataFrame(report_rows, columns=["Section", "Field", "Value"])

                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    report_df.to_excel(writer, index=False, sheet_name="Property Report")

                    worksheet = writer.sheets["Property Report"]

                    # Widen columns for readability
                    worksheet.column_dimensions["A"].width = 28
                    worksheet.column_dimensions["B"].width = 32
                    worksheet.column_dimensions["C"].width = 70

                    # Bold header row
                    from openpyxl.styles import Font, PatternFill
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = header_fill

                    # Shade each section's first row and bold the Section column
                    # to visually separate groups when scrolling
                    last_section = None
                    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    for row_idx in range(2, worksheet.max_row + 1):
                        section_cell = worksheet.cell(row=row_idx, column=1)
                        if section_cell.value != last_section:
                            for col in range(1, 4):
                                worksheet.cell(row=row_idx, column=col).fill = section_fill
                            section_cell.font = Font(bold=True)
                            last_section = section_cell.value

                buffer.seek(0)

                st.download_button(
                    label="📥 Download Full Report (Excel — structured)",
                    data=buffer.getvalue(),
                    file_name=f"risk_analysis_{uploaded_pdf.name}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    )
                )

            except Exception as e:
                st.error(f"❌ Error during analysis: {str(e)}")
            finally:
                os.unlink(tmp_path)

    # ===========================================================
    # MODE 2 — EXCEL BATCH
    # (unchanged — Excel batch mode is a separate, simpler flow
    #  using modules/extractor.py's analyze_property_row, which
    #  doesn't need property-type awareness since it scores
    #  whatever columns the user's Excel already has)
    # ===========================================================
    else:
        st.subheader("Step 2 — Upload Excel Property List")
        st.write(
            "Upload an Excel file — each row is one property. "
            "Columns can include: location, area, building_age, "
            "market_value, legal_status, road_access, etc."
        )

        with st.expander("📋 See expected Excel format"):
            sample_data = {
                "location": ["Koramangala", "Whitefield"],
                "area": [1200, 950],
                "building_age": [8, 15],
                "market_value": ["85,00,000", "62,00,000"],
                "legal_status": ["Clear", "Pending EC"],
                "road_access": ["30ft road", "20ft road"],
                "encumbrance": ["Clear", "Needs verification"]
            }
            st.dataframe(pd.DataFrame(sample_data))

        uploaded_excel = st.file_uploader(
            "📂 Upload Property Excel File",
            type=["xlsx"],
            key="risk_excel"
        )

        if uploaded_excel:
            df = pd.read_excel(uploaded_excel)
            st.subheader("📊 Input Data Preview")
            st.dataframe(df)

            if st.button(
                "🚀 Run Batch Risk Analysis", key="analyse_excel"
            ):
                results = []
                progress = st.progress(0)
                total = len(df)

                for i, row in df.iterrows():
                    property_data = row.to_dict()
                    st.write(
                        f"⚙ Analysing property {i + 1} of {total}..."
                    )

                    result = analyze_property_row(property_data)
                    validation = validate_property_data(property_data)

                    results.append({
                        **{k: v for k, v in property_data.items()},
                        "Risk Score": result.get("risk_score"),
                        "Investment Score": result.get("investment_score"),
                        "Market Liquidity": result.get("market_liquidity"),
                        "Price Range": result.get("price_range"),
                        "Key Risks": ", ".join(
                            result.get("key_risks", [])
                        ),
                        "Bank Ready": (
                            "✅ Yes" if validation["is_bank_ready"]
                            else "🔴 No"
                        ),
                        "Critical Issues": "; ".join(
                            f["field"]
                            for f in validation["critical_flags"]
                        ),
                        "Warnings": "; ".join(
                            f["field"]
                            for f in validation["warning_flags"]
                        ),
                    })

                    progress.progress((i + 1) / total)

                result_df = pd.DataFrame(results)

                try:
                    batch_id = save_excel_batch(
                        uploaded_excel.name,
                        result_df
                    )
                    st.write(
                        f"💾 Batch saved to database (ID: {batch_id})"
                    )
                except Exception as db_err:
                    st.warning(f"⚠️ Database save warning: {db_err}")

                st.session_state["risk_results"] = result_df
                st.success(
                    f"✅ Analysis complete — {total} properties scored!"
                )

                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total Properties", len(result_df))
                high_risk = sum(
                    result_df["Risk Score"].str.contains(
                        "High", na=False
                    )
                )
                medium_risk = sum(
                    result_df["Risk Score"].str.contains(
                        "Medium", na=False
                    )
                )
                not_ready = sum(
                    result_df["Bank Ready"].str.contains(
                        "No", na=False
                    )
                )
                col2.metric("🔴 High Risk", high_risk)
                col3.metric("🟡 Medium Risk", medium_risk)
                col4.metric("⛔ Not Bank Ready", not_ready)

                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("📈 Risk Distribution")
                    st.bar_chart(result_df["Risk Score"].value_counts())
                with col2:
                    st.subheader("✅ Bank Ready Status")
                    st.bar_chart(
                        result_df["Bank Ready"].value_counts()
                    )

                st.subheader("🏆 Top 5 Investment Properties")
                top = result_df.sort_values(
                    by="Investment Score", ascending=False
                ).head(5)
                st.dataframe(top)

                attention_df = result_df[
                    result_df["Bank Ready"].str.contains(
                        "No", na=False
                    )
                ]
                st.subheader("⛔ Properties Needing Attention")
                if len(attention_df) > 0:
                    st.dataframe(attention_df)
                else:
                    st.success("✅ All properties are bank-ready!")

                st.subheader("📊 Full Results")
                st.dataframe(result_df)

                buffer = BytesIO()
                result_df.to_excel(buffer, index=False)
                buffer.seek(0)
                st.download_button(
                    label="📥 Download Full Analysis (Excel)",
                    data=buffer.getvalue(),
                    file_name="property_batch_risk_analysis.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    )
                )

# ============================================================
# PAGE 5 — REPORTS
# ============================================================

elif page == "📋 Reports":
    st.title("📋 Reports")
    st.write(
        "Full history of all analysed properties — "
        "persists between sessions."
    )

    st.subheader("📄 PDF Valuation Report History")
    try:
        all_docs = get_all_documents()
        if all_docs:
            docs_df = pd.DataFrame(all_docs)
            st.dataframe(docs_df)

            buffer = BytesIO()
            docs_df.to_excel(buffer, index=False)
            buffer.seek(0)
            st.download_button(
                label="📥 Download Document History (Excel)",
                data=buffer.getvalue(),
                file_name="document_history.xlsx",
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                )
            )
        else:
            st.info(
                "No PDF analyses saved yet. "
                "Go to 📊 Risk Analysis → PDF mode to analyse a report."
            )
    except Exception as e:
        st.error(f"❌ Could not load document history: {str(e)}")

    st.divider()

    st.subheader("📊 Excel Batch Analysis History")
    try:
        all_batches = get_all_excel_batches()
        if all_batches:
            batches_df = pd.DataFrame(all_batches)
            st.dataframe(batches_df)

            buffer = BytesIO()
            batches_df.to_excel(buffer, index=False)
            buffer.seek(0)
            st.download_button(
                label="📥 Download Batch History (Excel)",
                data=buffer.getvalue(),
                file_name="batch_history.xlsx",
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                )
            )
        else:
            st.info(
                "No Excel batches saved yet. "
                "Go to 📊 Risk Analysis → Excel mode to run a batch."
            )
    except Exception as e:
        st.error(f"❌ Could not load batch history: {str(e)}")

    st.divider()

    if "risk_results" in st.session_state:
        st.subheader("📊 Current Session — Excel Results")
        st.dataframe(st.session_state["risk_results"])